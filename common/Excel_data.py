from openpyxl import load_workbook

class Case:
    def __init__(self):
        self.case_id=None
        self.title=None
        self.url=None
        self.data=None
        self.method=None
        self.expected=None
        self.sql=None

class Do_Excel:
    def __init__(self,filename,sheet_name):
        self.filename=filename
        self.fn = load_workbook(self.filename)        #打开文件
        self.sheet_name=sheet_name
        self.sheet=self.fn[self.sheet_name]

    def read_date(self):
        list1 = []  # 创建大列表
        max_row = self.sheet.max_row
        max_column = self.sheet.max_column
        #  定位到表单
        for i in range(2,max_row + 1):  # 定位单元格的行
            list = []  # 小列表   每次重新赋值该列表 始终保证每一行为一个列表
            for j in range(1,max_column + 1):  # 定位单元格的列
                if self.sheet.cell(i, j).value:  # 判断是否有None
                    list.append(self.sheet.cell(i, j).value)  # 插入list列表
            list1.append(list)  # 小列表插入大列表
        return list1  # 返回值

    def read_Excel(self):
        max_row = self.sheet.max_row  # 获取最大行数
        cases = []  # 列表，存放所有的测试用例
        for r in range(2, max_row + 1):
            case = Case()  # 实例
            case.case_id = self.sheet.cell(row=r, column=1).value
            case.title = self.sheet.cell(row=r, column=2).value
            case.url = self.sheet.cell(row=r, column=3).value
            case.data = self.sheet.cell(row=r, column=4).value
            case.method = self.sheet.cell(row=r, column=5).value
            case.expected = self.sheet.cell(row=r, column=6).value
            case.sql = self.sheet.cell(row=r, column=9).value
            cases.append(case)
        self.fn.close()
        return cases  # 返回case列表

    def write_excel(self,row,actual,result):
        self.sheet.cell(row,7).value = actual
        self.sheet.cell(row,8).value = result
        self.fn.save(filename=self.filename)
        self.fn.close()

if __name__ == '__main__':
 r=Do_Excel('D:\\test_dentalink\\trunk\\Test_dental\\data\\cases2.xlsx','patsplitsub').read_Excel()
 s= eval(r[0].data)
 print(type(s))


